"""
ADR Tablo A (ADR_A_TABLOSU.xlsx) üzerinde UN No + Sınıf + Paketleme Grubu
ile tam eşleştirme yapan modül.
"""
import openpyxl

NOT_IN_SCOPE_TEXT = "SDS RAPORU BÖLÜM 14 KONTROLÜNE İSTİNADEN ÜRÜN ADR KAPSAMINDA DEĞİLDİR"
MANUAL_REVIEW_TEXT = "MANUEL KONTROL GEREKLİ"

_cache = {}


def load_tablo_a(path: str):
    if path in _cache:
        return _cache[path]
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["ADR A TABLOSU"]
    rows = []
    for r in range(5, ws.max_row + 1):
        un = ws.cell(row=r, column=1).value
        if un is None:
            continue
        rows.append({
            "un_no": str(un).strip(),
            "isim": ws.cell(row=r, column=2).value,
            "sinif": str(ws.cell(row=r, column=3).value).strip() if ws.cell(row=r, column=3).value is not None else None,
            "siniflandirma_kodu": (
                str(ws.cell(row=r, column=4).value).strip()
                if ws.cell(row=r, column=4).value is not None
                else None
            ),
            "paketleme_grubu": (ws.cell(row=r, column=5).value or "").strip() or None,
            "ozel_hukumler": ws.cell(row=r, column=7).value,
            "sinirli_miktar": ws.cell(row=r, column=8).value,
            "istisnai_miktar": ws.cell(row=r, column=9).value,
            "paketleme_talimati": ws.cell(row=r, column=10).value,
            "tank_kodu": ws.cell(row=r, column=15).value,
            "tasima_kategori": ws.cell(row=r, column=18).value,
        })
    _cache[path] = rows
    return rows


def match_tablo_a(
    tablo_a_path: str,
    un_no: str,
    sinif: str,
    paketleme_grubu: str,
    siniflandirma_kodu: str = None,
):
    """
    Öncelik sırası

    1) UN + SINIF + PG

    2) Eğer PG yoksa

       UN + SINIF + SINIFLANDIRMA KODU
    """

    rows = load_tablo_a(tablo_a_path)

    un_no = str(un_no).strip()
    sinif = str(sinif).strip()

    pg = (paketleme_grubu or "").strip() or None

    sk = (
        str(siniflandirma_kodu).strip()
        if siniflandirma_kodu
        else None
    )

    # --------------------------------------------------
    # 1) Önce PG ile tam eşleşme
    # --------------------------------------------------

    for row in rows:

        if (
            row["un_no"] == un_no
            and row["sinif"] == sinif
            and row["paketleme_grubu"] == pg
        ):
            return row

    # --------------------------------------------------
    # Aynı UN + sınıf kayıtlarını topla
    # --------------------------------------------------

    same_rows = []

    for row in rows:

        if (
            row["un_no"] == un_no
            and row["sinif"] == sinif
        ):
            same_rows.append(row)

    if not same_rows:
        return None

    # --------------------------------------------------
    # Eğer bu UN+sınıfta PG kullanan kayıt varsa
    # classification code ile arama yapma
    # --------------------------------------------------

    has_pg = any(
        r["paketleme_grubu"]
        for r in same_rows
    )

    if has_pg:
        return None

    # --------------------------------------------------
    # PG gerçekten yok (Sınıf 2, Sınıf 7, bazı Sınıf 9 vb.)
    # Akış:
    #   1) Tablo A'da PG yok mu? ✓ (buraya geldik, has_pg=False)
    #   2) PDF'ten siniflandirma_kodu geldi mi?
    #      → Geldiyse: kod eşleştirmesini dene
    #      → Eşleştiyse: döndür
    #   3) Siniflandirma kodu yoksa VEYA eşleşmediyse:
    #      → Tek satır varsa: döndür (zaten başka ihtimal yok)
    #      → Birden fazla satır varsa: MANUEL KONTROL
    #        (hangi siniflandirma kodunun doğru olduğunu bilemeyiz)
    # --------------------------------------------------

    # Adım 2: PDF'ten gelen siniflandirma_kodu ile eşleştirmeyi dene
    if sk:
        for row in same_rows:
            row_sk = (
                str(row["siniflandirma_kodu"]).strip()
                if row["siniflandirma_kodu"]
                else None
            )
            if row_sk == sk:
                return row

    # Adım 3: Siniflandirma kodu yoksa veya eşleşmediyse
    if len(same_rows) == 1:
        return same_rows[0]

    # Birden fazla satır var, siniflandirma kodu da yok/eşleşmedi:
    # MANUEL KONTROL -- hangi satırın doğru olduğunu bilemeyiz.
    return None


def get_official_sinif_for_un(tablo_a_path: str, un_no: str):
    """Sadece UN numarasına bakarak (sınıf/paketleme grubundan bağımsız)
    Tablo A'daki RESMİ sınıfı döndürür. PDF'i hazırlayanların sınıfı
    yanlış yazmış olabileceği durumları tespit etmek için kullanılır
    (örn. PDF '9' diyor ama o UN no resmen sadece '6.1' altında yer alır)."""
    rows = load_tablo_a(tablo_a_path)
    for row in rows:
        if row["un_no"] == str(un_no).strip():
            return row["sinif"]
    return None


def build_inventory_row(adr_info: dict, tablo_a_path: str, kimyasal_adi: str,
                         ambalaj_tank_dokme: str = "AMBALAJLI"):
    """extractor.extract_adr_info() çıktısından envanter satırı sözlüğü üretir."""
    row = {
        "Kimyasal Adı": kimyasal_adi,
        "AMBALAJLI/TANK/DÖKME": ambalaj_tank_dokme,
        "MSDS/SDS TARİHİ": adr_info.get("revize_tarihi"),
        "ADR-IMDG-IATA": "ADR",  # Bu sütun her zaman sabit "ADR" yazar (durum ne olursa olsun)
        "Cas_No": adr_info.get("cas_no"),
        "Tedarikçi": adr_info.get("tedarikci"),
        "Fonksiyonu": adr_info.get("fonksiyon"),
        "Tehlikeli/ Tehlikesiz": adr_info.get("tehlikeli_tehlikesiz"),
        "H KODLARI": adr_info.get("h_kodlari"),
        "durum": "ok",  # ok | not_in_scope | manual_review
    }
    # "Tehlike Etiketi" sütunu görsel (piktogram) için ayrılmıştır, metin
    # yazılmaz -- UN no gerçekten varsa (bu fonksiyonun en altındaki "ok"
    # yolu) hiç dokunulmaz; kapsam dışı/manuel kontrolde ise sadece ürün
    # KESİN tehlikesizse "-" yazılır (Versiyon 2 ile aynı kural).

    if adr_info.get("adr_kapsaminda") is False:
        row.update({
            "UN NUMARASI": NOT_IN_SCOPE_TEXT,
            "SINIFI / ETİKETİ": NOT_IN_SCOPE_TEXT,
            "PAKETLEME GRUBU": NOT_IN_SCOPE_TEXT,
            "UYGUN SEVKİYAT ADI": NOT_IN_SCOPE_TEXT,
            "SINIRLI MİKTAR": NOT_IN_SCOPE_TEXT,
            "İSTİSNAİ MİKTAR": NOT_IN_SCOPE_TEXT,
            "ÖZEL HÜKÜMLER": NOT_IN_SCOPE_TEXT,
            "TANK KODU": NOT_IN_SCOPE_TEXT,
            "AMBALAJLAMA TALİMATLARI": NOT_IN_SCOPE_TEXT,
            "TAŞIMA KATEGORİSİ/(TÜNEL KODU)": NOT_IN_SCOPE_TEXT,
        })
        if adr_info.get("tehlikeli_tehlikesiz") == "Tehlikesiz":
            row["Tehlike Etiketi"] = "-"
        row["durum"] = "not_in_scope"
        return row

    if adr_info.get("adr_kapsaminda") is None or not adr_info.get("un_no"):
        row.update({
            "UN NUMARASI": MANUAL_REVIEW_TEXT,
            "SINIFI / ETİKETİ": MANUAL_REVIEW_TEXT,
            "PAKETLEME GRUBU": MANUAL_REVIEW_TEXT,
            "UYGUN SEVKİYAT ADI": MANUAL_REVIEW_TEXT,
            "SINIRLI MİKTAR": MANUAL_REVIEW_TEXT,
            "İSTİSNAİ MİKTAR": MANUAL_REVIEW_TEXT,
            "ÖZEL HÜKÜMLER": MANUAL_REVIEW_TEXT,
            "TANK KODU": MANUAL_REVIEW_TEXT,
            "AMBALAJLAMA TALİMATLARI": MANUAL_REVIEW_TEXT,
            "TAŞIMA KATEGORİSİ/(TÜNEL KODU)": MANUAL_REVIEW_TEXT,
        })
        if adr_info.get("tehlikeli_tehlikesiz") == "Tehlikesiz":
            row["Tehlike Etiketi"] = "-"
        row["durum"] = "manual_review"
        return row

    un_no, sinif, pg = (
        adr_info["un_no"],
        adr_info["sinif"],
        adr_info["paketleme_grubu"],
    )

    match = match_tablo_a(
        tablo_a_path,
        un_no,
        sinif,
        pg,
        adr_info.get("siniflandirma_kodu"),
)

    duzeltme_notu = None
    if match is None:
        # UN no Tablo A'da var ama PDF'teki sınıf/PG kombinasyonu eşleşmiyor
        # olabilir. İki ayrı durum var:
        #   a) PDF'te sınıf hiç okunamadı (sinif=None/boş) -- bu MSDS'in
        #      hatası değil, çıkarma işleminin o PDF'te sınıfı bulamamış
        #      olmasıdır. Üreticiye bildirme önerisi YANLIŞ olur.
        #   b) PDF'te GERÇEK ama YANLIŞ bir sınıf yazılı (örn. '9' yazılı
        #      ama resmi sınıf '6.1') -- bu durumda MSDS'i hazırlayan
        #      firmanın hatası olabilir, bildirilmesi mantıklı.
        official_sinif = get_official_sinif_for_un(tablo_a_path, un_no)
        sinif_bos = not str(sinif).strip() or str(sinif).strip().lower() == "none"
        if official_sinif and (sinif_bos or official_sinif != str(sinif).strip()):
            duzeltilmis_match = match_tablo_a(
                tablo_a_path,
                un_no,
                official_sinif,
                pg,
                adr_info.get("siniflandirma_kodu"),
            )
            if duzeltilmis_match:
                match = duzeltilmis_match
                if sinif_bos:
                    duzeltme_notu = (
                        f"ℹ️ TAMAMLANDI: MSDS'te UN {un_no} için Bölüm 14'te Sınıf "
                        f"bilgisi okunamadı; ADR Tablo A'ya göre bu UN'nin resmi "
                        f"sınıfı ({official_sinif}) kullanılarak dolduruldu."
                    )
                else:
                    duzeltme_notu = (
                        f"⚠️ DÜZELTİLDİ: MSDS'te UN {un_no} için Sınıf {sinif} yazılıydı; "
                        f"ADR Tablo A'ya göre resmi sınıf {official_sinif} olduğu için bu "
                        f"değer kullanılarak dolduruldu. MSDS hazırlayan firmaya bildirin."
                    )
                sinif = official_sinif

    row["UN NUMARASI"] = int(un_no)
    row["SINIFI / ETİKETİ"] = int(sinif) if str(sinif).isdigit() else sinif
    row["ADR-IMDG-IATA"] = "ADR"

    if match is None:
        row["PAKETLEME GRUBU"] = pg or MANUAL_REVIEW_TEXT
    else:
        # Eşleşme bulunduysa, Paketleme Grubu'nu PDF'ten gelen (bazen
        # tutarsız/hatalı olabilen) değer yerine DOĞRUDAN Tablo A'nın
        # eşleşen satırından çekiyoruz -- diğer tüm alanlar (Uygun
        # Sevkiyat Adı, Tank Kodu vb.) da zaten aynı şekilde Tablo A'dan
        # geliyor, tutarlı olsun. Tablo A'da bu UN/sınıf için PG yoksa
        # (örn. gazlar, Sınıf 1) bu None'dır ve hücre boş bırakılır --
        # "MANUEL KONTROL GEREKLİ" YAZILMAZ, çünkü bu bir hata değildir.
        row["PAKETLEME GRUBU"] = match["paketleme_grubu"]

    if match is None:
        row.update({
            "UYGUN SEVKİYAT ADI": MANUAL_REVIEW_TEXT,
            "SINIRLI MİKTAR": MANUAL_REVIEW_TEXT,
            "İSTİSNAİ MİKTAR": MANUAL_REVIEW_TEXT,
            "ÖZEL HÜKÜMLER": MANUAL_REVIEW_TEXT,
            "TANK KODU": MANUAL_REVIEW_TEXT,
            "AMBALAJLAMA TALİMATLARI": MANUAL_REVIEW_TEXT,
            "TAŞIMA KATEGORİSİ/(TÜNEL KODU)": MANUAL_REVIEW_TEXT,
        })
        # Resmi sınıfla bile eşleşme bulunamadıysa (örn. PG de yanlış
        # olabilir), en azından tespit edilen sınıf tutarsızlığını bildir.
        official_sinif = get_official_sinif_for_un(tablo_a_path, un_no)
        sinif_bos = not str(sinif).strip() or str(sinif).strip().lower() == "none"
        if official_sinif and (sinif_bos or official_sinif != str(sinif).strip()):
            if sinif_bos:
                row["Açıklama"] = (
                    f"ℹ️ MSDS'te UN {un_no} için Bölüm 14'te Sınıf bilgisi okunamadı. "
                    f"ADR Tablo A'ya göre resmi sınıfı {official_sinif}, ancak Paketleme "
                    f"Grubu uyuşmadığı için satır elle kontrol edilmelidir."
                )
            else:
                row["Açıklama"] = (
                    f"⚠️ UYARI: MSDS'te UN {un_no} için Sınıf {sinif} yazılı, ancak ADR "
                    f"Tablo A'ya göre bu UN numarasının resmi sınıfı {official_sinif}. "
                    f"MSDS hazırlayan firmaya danışıp elle düzeltin."
                )
        row["durum"] = "manual_review"
        return row

    if duzeltme_notu:
        row["Açıklama"] = duzeltme_notu

    row["UYGUN SEVKİYAT ADI"] = match["isim"]
    row["SINIRLI MİKTAR"] = match["sinirli_miktar"]
    row["İSTİSNAİ MİKTAR"] = match["istisnai_miktar"]
    row["ÖZEL HÜKÜMLER"] = match["ozel_hukumler"]
    row["TANK KODU"] = match["tank_kodu"]
    row["AMBALAJLAMA TALİMATLARI"] = match["paketleme_talimati"]
    row["TAŞIMA KATEGORİSİ/(TÜNEL KODU)"] = match["tasima_kategori"]
    return row


NOT_IN_SCOPE_TEXT_V2 = "MSDS/SDS Raporu bölüm 14 kapsamında ADR kapsamında değildir."


def build_inventory_row_v2(adr_info: dict, kimyasal_adi: str, ambalaj_tank_dokme: str = None):
    """ORDU tarzı (Versiyon 2) basit envanter formatı için satır sözlüğü
    üretir. V1'den farkı: Paketleme Grubu, Sınırlı Miktar, Tank Kodu vb.
    ayrıntılı Tablo A sütunları yok -- sadece UN No, Sınıf ve ADR İşareti.
    Bu yüzden Tablo A ile eşleştirme yapılmaz, sadece PDF'ten okunan ham
    bilgi yazılır.

    "Tehlike_Etiketi" ve "ADR_İŞARETİ" sütunları görsel (piktogram) için
    ayrılmıştır, metin yazılmaz:
      - O satırda gerçek bir UN no varsa (ADR kapsamında): ikisi de
        TAMAMEN ATLANIR (piktogram elle eklenecek).
      - UN no yoksa (kapsam dışı/manuel kontrol): ADR_İŞARETİ yine de
        durum metnini alır (önceden onaylandı); Tehlike_Etiketi'ne ise
        sadece ürün KESİN tehlikesizse "-" yazılır, tehlikeliyse (H kodu
        varsa) piktogram elle eklenecek şekilde BOŞ bırakılır.
    """
    row = {
        "Kimyasal Adı": kimyasal_adi,
        "MSDS/SDS TARİHİ": adr_info.get("revize_tarihi"),
        "Sistem_Kodu/Tedarikçi Firma": adr_info.get("tedarikci"),
        "Cas_No": adr_info.get("cas_no"),
        "H KODLARI": adr_info.get("h_kodlari"),
        "Fonksiyonu": adr_info.get("fonksiyon"),
        "Tehlikeli/Tehlikesiz": adr_info.get("tehlikeli_tehlikesiz"),
        "durum": "ok",
    }
    if ambalaj_tank_dokme:
        row["Ambalaj/Tank"] = ambalaj_tank_dokme

    un_no_var = bool(adr_info.get("adr_kapsaminda") is True and adr_info.get("un_no"))

    if adr_info.get("adr_kapsaminda") is False:
        row.update({
            "UN NUMARASI": NOT_IN_SCOPE_TEXT_V2,
            "SINIFI": NOT_IN_SCOPE_TEXT_V2,
            "ADR_İŞARETİ": NOT_IN_SCOPE_TEXT_V2,
        })
        if adr_info.get("tehlikeli_tehlikesiz") == "Tehlikesiz":
            row["Tehlike_Etiketi"] = "-"
        # "Tehlikeli" veya belirsizse Tehlike_Etiketi'ne hiç dokunulmaz (piktogram elle eklenecek).
        row["durum"] = "not_in_scope"
        return row

    if adr_info.get("adr_kapsaminda") is None or not adr_info.get("un_no"):
        row.update({
            "UN NUMARASI": MANUAL_REVIEW_TEXT,
            "SINIFI": MANUAL_REVIEW_TEXT,
            "ADR_İŞARETİ": MANUAL_REVIEW_TEXT,
        })
        if adr_info.get("tehlikeli_tehlikesiz") == "Tehlikesiz":
            row["Tehlike_Etiketi"] = "-"
        row["durum"] = "manual_review"
        return row

    row["UN NUMARASI"] = f"UN {adr_info['un_no']}"
    row["SINIFI"] = adr_info.get("sinif")
    return row


if __name__ == "__main__":
    from extractor import extract_adr_info
    import json

    tablo_a = "ADR_A_TABLOSU.xlsx"
    tests = [
        ("ADVANTAGE_101M_MSDS_TÜRKÇE.PDF", "ADVANTAGE 101M"),
        ("ADVANTAGE_121_ODT_MSDS_TÜRKÇE.pdf", "ADVANTAGE 121 ODT"),
    ]
    for pdf, ad in tests:
        info = extract_adr_info(pdf)
        row = build_inventory_row(info, tablo_a, ad)
        print("=" * 80)
        print(pdf)
        print(json.dumps(row, ensure_ascii=False, indent=2))

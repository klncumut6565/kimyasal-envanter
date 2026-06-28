import io
import os
import re
import tempfile
import time

import streamlit as st
from openpyxl import load_workbook

from extractor import extract_adr_info, clean_product_name
from matcher import (build_inventory_row, build_inventory_row_v2,
                       NOT_IN_SCOPE_TEXT, MANUAL_REVIEW_TEXT)
from excel_writer import (add_products, fill_or_append_v2, create_new_envanter,
                            find_column, SHEET_NAME, HEADER_ROW)

st.set_page_config(page_title="Kimyasal Envanter Oluşturucu", layout="wide")

# ADR Tablo A, programla birlikte gelen sabit bir referans dosyasıdır.
# Kullanıcı bunu her seferinde yüklemek zorunda değildir.
TABLO_A_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data", "ADR_A_TABLOSU.xlsx")
BOS_SABLON_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data", "BOS_ENVANTER_SABLONU.xlsx")

if "tmp_dir" not in st.session_state:
    st.session_state.tmp_dir = tempfile.mkdtemp(prefix="kimyasal_envanter_")
if "urunler" not in st.session_state:
    st.session_state.urunler = {}  # key: pdf dosya adı -> dict
if "sonuc_dosyasi" not in st.session_state:
    st.session_state.sonuc_dosyasi = None
if "sonuc_mesajlari" not in st.session_state:
    st.session_state.sonuc_mesajlari = []

TMP = st.session_state.tmp_dir


def save_upload(uploaded_file, subdir=""):
    d = os.path.join(TMP, subdir)
    os.makedirs(d, exist_ok=True)
    path = os.path.join(d, uploaded_file.name)
    with open(path, "wb") as f:
        f.write(uploaded_file.getbuffer())
    return path


def existing_names(envanter_path):
    try:
        wb = load_workbook(envanter_path, data_only=True)
        ws = wb[SHEET_NAME]
        ad_col = find_column(ws, "Kimyasal Adı") or 2
        names = set()
        for r in range(HEADER_ROW + 1, ws.max_row + 1):
            v = ws.cell(row=r, column=ad_col).value
            if isinstance(v, str) and v.strip():
                names.add(v.strip().upper())
        return names
    except Exception:
        return set()


def render_export_ui(secili_urunler, envanter_path, v2, firma_adi, key_suffix):
    """'Excel'e Aktar' butonu + sonuç mesajları + indirme butonu. Hem sayfanın
    üstünde hem altında çağrılır ki kullanıcı 500+ ürün varken sayfa sonunu
    aramak zorunda kalmasın."""
    st.write(f"**{len(secili_urunler)} ürün** Excel'e eklenmeye hazır.")

    # Çıktı dosya adına firma adını ekle (V1 için) -- birden çok firma ile
    # çalışırken indirilen dosyalar birbirine karışmasın.
    firma_temiz = re.sub(r'[\\/*?:"<>|]', "", (firma_adi or "")).strip().replace(" ", "_")
    dosya_adi = f"{firma_temiz}_envanter_guncellendi.xlsx" if (firma_temiz and not v2) else "envanter_guncellendi.xlsx"

    if st.button("📥 Tüm Seçili Ürünleri Excel'e Aktar", type="primary",
                  disabled=len(secili_urunler) == 0, key=f"export_btn_{key_suffix}"):
        rows = []
        for u in secili_urunler:
            if v2:
                row = build_inventory_row_v2(u["info"], u["kimyasal_adi"], u["ambalaj"])
            else:
                row = build_inventory_row(u["info"], TABLO_A_PATH, u["kimyasal_adi"], u["ambalaj"])
            rows.append(row)

        out_path = os.path.join(TMP, dosya_adi)
        st.session_state.sonuc_mesajlari = []
        if v2:
            try:
                sonuc = fill_or_append_v2(envanter_path, out_path, rows)
            except ValueError as e:
                sonuc = None
                st.session_state.sonuc_mesajlari.append(("error", f"❌ {e}"))
            if sonuc is not None:
                dolan = [ad for _, d, ad in sonuc if d == "dolduruldu"]
                zaten_dolu = [ad for _, d, ad in sonuc if d == "zaten_dolu"]
                eslesmeyen = [ad for _, d, ad in sonuc if d == "eslesme_yok"]
                st.session_state.sonuc_dosyasi = out_path
                if dolan:
                    st.session_state.sonuc_mesajlari.append(
                        ("success", f"✅ {len(dolan)} ürünün boş hücreleri dolduruldu: {', '.join(dolan)}"))
                if zaten_dolu:
                    st.session_state.sonuc_mesajlari.append(
                        ("info", f"ℹ️ {len(zaten_dolu)} ürün eşleşti ama hücreleri zaten doluydu, "
                                 f"değişiklik yapılmadı: {', '.join(zaten_dolu)}"))
                if eslesmeyen:
                    st.session_state.sonuc_mesajlari.append(
                        ("warning", f"⚠️ {len(eslesmeyen)} ürün envanterde bulunamadığı için ATLANDI "
                                    f"(Versiyon 2 yeni satır eklemez): {', '.join(eslesmeyen)}"))
        else:
            added = add_products(envanter_path, out_path, rows)
            st.session_state.sonuc_dosyasi = out_path
            st.session_state.sonuc_mesajlari.append(
                ("success", f"{len(added)} yeni satır eklendi (Excel satırları: {added})."))

    for tip, mesaj in st.session_state.sonuc_mesajlari:
        getattr(st, tip)(mesaj)

    if st.session_state.sonuc_dosyasi and os.path.exists(st.session_state.sonuc_dosyasi):
        with open(st.session_state.sonuc_dosyasi, "rb") as f:
            st.download_button(
                "⬇️ Güncellenmiş Envanter Excel Dosyasını İndir",
                data=f.read(),
                file_name=dosya_adi,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key=f"download_btn_{key_suffix}",
            )


st.title("🧪 Kimyasal Envanter Oluşturucu")
st.caption("MSDS PDF → Bölüm 14 (Taşıma Bilgileri) → ADR Tablo A eşleştirme → Envanter Excel")

with st.sidebar:
    st.header("0) Şablon Versiyonu")
    versiyon = st.radio(
        "Envanter dosyanızın formatı hangisi?",
        ["Versiyon 1 (Paketleme Grubu, Sınırlı Miktar vb. ayrıntılı sütunlar)",
         "Versiyon 2 (basit format — sadece UN No / Sınıf / ADR İşareti)"],
        key="versiyon",
    )
    v2 = versiyon.startswith("Versiyon 2")

    st.header("1) Envanter Dosyası")

    if not os.path.exists(TABLO_A_PATH) and not v2:
        st.error(
            "ADR Tablo A referans dosyası programda bulunamadı "
            f"({TABLO_A_PATH}). Lütfen geliştiriciyle iletişime geçin."
        )

    envanter_path = None
    firma_adi = ""

    if v2:
        st.caption("Versiyon 2'de program, ürün adı zaten satırda varsa "
                    "boş ADR hücrelerini doldurur (ADR İşareti sütununa hiç dokunmaz); "
                    "eşleşme bulunamazsa o ürün atlanır, yeni satır eklenmez. "
                    "Dosyada **'KİMYASAL ENVANTER2026'** adlı bir sayfa olmalı.")
        envanter_file = st.file_uploader(
            "Versiyon 2 formatındaki Envanter Excel Dosyası (.xlsx)",
            type=["xlsx"], key="envanter_v2")
        envanter_path = save_upload(envanter_file) if envanter_file else None
        if envanter_path:
            st.success("Envanter dosyası yüklendi ✓")
        else:
            st.info("Devam etmek için envanter Excel dosyasını yükleyin.")
    else:
        mod = st.radio(
            "Nasıl başlamak istersiniz?",
            ["🆕 Yeni envanter oluştur", "📂 Var olan envanter çıktısını güncelle"],
            key="mod",
        )

        if mod.startswith("🆕"):
            st.caption("ADR Tablo A ve tüm başlık/imza biçimi programda hazır gelir; "
                       "sadece firma bilgilerinizi girmeniz yeterli.")
            firma_adi = st.text_input("Firma Adı", key="firma_adi",
                                       placeholder="Örn. ASUTEK")
            hazirlayan_adi = st.text_input("Hazırlayan Adı (Tehlikeli Madde Güvenlik Danışmanı)",
                                            key="hazirlayan_adi", placeholder="Örn. Ahmet Yılmaz")
            onaylayan_adi = st.text_input("Onaylayan Adı Soyadı (boş bırakılırsa \"Sorumlu Kişi\" yazılır)",
                                           key="onaylayan_adi", placeholder="Örn. Mehmet Demir")
            firma_logo_file = st.file_uploader(
                "Firma Logosu (opsiyonel, .png/.jpg)", type=["png", "jpg", "jpeg"], key="firma_logo")

            if firma_adi.strip():
                firma_logo_path = save_upload(firma_logo_file, subdir="firma_logo") if firma_logo_file else None
                son_anahtar = (firma_adi.strip(), hazirlayan_adi.strip(), onaylayan_adi.strip(),
                               firma_logo_file.name if firma_logo_file else None)
                if st.session_state.get("yeni_envanter_anahtar") != son_anahtar:
                    envanter_path = os.path.join(TMP, "yeni_envanter.xlsx")
                    create_new_envanter(BOS_SABLON_PATH, envanter_path, firma_adi,
                                         firma_logo_path, hazirlayan_adi, onaylayan_adi)
                    st.session_state.yeni_envanter_anahtar = son_anahtar
                    st.session_state.yeni_envanter_path = envanter_path
                envanter_path = st.session_state.yeni_envanter_path
                st.success("Yeni envanter şablonu hazır ✓")
            else:
                st.info("Devam etmek için firma adını girin.")
        else:
            firma_adi = st.text_input(
                "Firma Adı (sadece indirilecek dosya adında kullanılır, karışıklığı önlemek için)",
                key="firma_adi_existing", placeholder="Örn. ASUTEK")
            envanter_file = st.file_uploader(
                "Bu programla daha önce oluşturulmuş Envanter Excel Dosyası (.xlsx)",
                type=["xlsx"], key="envanter")
            envanter_path = save_upload(envanter_file) if envanter_file else None
            if envanter_path:
                st.success("Envanter dosyası yüklendi ✓")
            else:
                st.info("Devam etmek için envanter Excel dosyasını yükleyin.")

st.header("2) MSDS PDF'lerini Yükle")
pdf_files = st.file_uploader(
    "Bir veya birden çok MSDS PDF dosyası seçin",
    type=["pdf"], accept_multiple_files=True, key="pdfs",
)

mevcut_isimler = existing_names(envanter_path) if (envanter_path and not v2) else set()

tablo_a_hazir = v2 or os.path.exists(TABLO_A_PATH)

if pdf_files and envanter_path and tablo_a_hazir:
    yeni_pdfler = [p for p in pdf_files if p.name not in st.session_state.urunler]

    if yeni_pdfler:
        toplam = len(yeni_pdfler)
        ilerleme_metni = st.empty()
        ilerleme_cubugu = st.progress(0.0)
        sure_metni = st.empty()
        sure_gecmisleri = []  # her PDF'in işlem süresi (saniye)

        for i, pdf in enumerate(yeni_pdfler, start=1):
            ilerleme_metni.write(f"📄 İşleniyor: **{pdf.name}** ({i}/{toplam})")
            pdf_path = save_upload(pdf, subdir="pdf")
            _baslangic = time.time()
            try:
                info = extract_adr_info(pdf_path)
            except Exception as e:
                # Bozuk/okunamayan bir PDF tüm toplu işlemi durdurmasın --
                # bu ürün "manuel kontrol gerekli" olarak işaretlenir,
                # diğer PDF'lerin işlenmesine devam edilir.
                info = {
                    "revize_tarihi": None, "onerilen_ad": None, "un_no": None,
                    "sinif": None, "paketleme_grubu": None, "adr_kapsaminda": None,
                    "ham_metin_bulundu": False, "tedarikci": None, "fonksiyon": None,
                    "cas_no": None, "h_kodlari": None, "tehlikeli_tehlikesiz": None,
                    "tehlike_etiketi": None, "okuma_hatasi": str(e),
                }
            # Kimyasal Adı önce PDF İÇERİĞİNDEN aranır; içerikte bulunamazsa
            # PDF dosya adına geri dönülür -- ama clean_product_name() bu
            # dosya adındaki "MSDS", "TR-SDS", "rev 7" gibi gerçek ürün adı
            # olmayan ekleri temizler, ham dosya adı asla yazılmaz.
            onerilen_ad = clean_product_name(info.get("onerilen_ad") or os.path.splitext(pdf.name)[0])
            st.session_state.urunler[pdf.name] = {
                "pdf_path": pdf_path,
                "info": info,
                "kimyasal_adi": onerilen_ad,
                "ambalaj": "Ambalaj" if v2 else "AMBALAJLI",
                "logo_path": None,
                "dahil_et": True,
            }
            ilerleme_cubugu.progress(i / toplam)

            # Geçen süreyi kaydet ve tahmini kalan süreyi hesapla
            sure_gecmisleri.append(time.time() - _baslangic)
            kalan_pdf = toplam - i
            if kalan_pdf > 0:
                ort_sure = sum(sure_gecmisleri) / len(sure_gecmisleri)
                tahmini_kalan = ort_sure * kalan_pdf
                if tahmini_kalan < 60:
                    sure_metni.info(f"⏱️ {kalan_pdf} PDF kaldı — tahmini kalan süre: **~{tahmini_kalan:.0f} sn**")
                else:
                    dk = int(tahmini_kalan // 60)
                    sn = int(tahmini_kalan % 60)
                    sure_metni.info(f"⏱️ {kalan_pdf} PDF kaldı — tahmini kalan süre: **~{dk} dk {sn} sn**")
            else:
                sure_metni.empty()

        ilerleme_metni.empty()
        ilerleme_cubugu.empty()
        sure_metni.empty()
        st.toast(f"✅ {toplam} PDF işlendi", icon="✅")

    st.divider()
    secili_urunler = [u for u in st.session_state.urunler.values() if u["dahil_et"]]
    st.subheader("⬆️ Hızlı Aktarım (sayfa sonuna inmenize gerek yok)")
    render_export_ui(secili_urunler, envanter_path, v2, firma_adi, key_suffix="top")
    st.divider()

    st.header("3) Çıkarılan Bilgileri Gözden Geçirin")

    for fname, urun in st.session_state.urunler.items():
        info = urun["info"]
        durum = (
            "🔴 PDF OKUNAMADI (dosya bozuk olabilir)" if info.get("okuma_hatasi")
            else "🟢 ADR kapsamında - Tablo A eşleşmesi bulundu" if info.get("adr_kapsaminda") and info.get("un_no")
            else "⚪ ADR kapsamında değil" if info.get("adr_kapsaminda") is False
            else "🟠 MANUEL KONTROL GEREKLİ (Bölüm 14 otomatik okunamadı)"
        )

        with st.expander(f"📄 {fname} — {durum}", expanded=True):
            if info.get("okuma_hatasi"):
                st.error(
                    "Bu PDF okunamadı, dosya bozuk veya standart olmayan bir yapıda olabilir. "
                    "Tüm bilgileri elle girmeniz gerekecek. "
                    f"(Teknik detay: {info['okuma_hatasi'][:200]})"
                )
            col1, col2 = st.columns([2, 1])
            with col1:
                urun["kimyasal_adi"] = st.text_input(
                    "Kimyasal Adı (Envantere yazılacak isim)",
                    value=urun["kimyasal_adi"], key=f"ad_{fname}",
                )
                if v2:
                    if urun["kimyasal_adi"].strip():
                        st.info(
                            f"ℹ️ Bu isim envanterde bulunursa boş hücreleri dolduracak; "
                            "bulunamazsa bu ürün ATLANACAK (Versiyon 2 yeni satır eklemez)."
                        )
                elif urun["kimyasal_adi"].strip().upper() in mevcut_isimler:
                    st.warning(
                        f"⚠️ '{urun['kimyasal_adi']}' adı envanterde zaten mevcut! "
                        "Mükerrer satır oluşturmamak için isimi kontrol edin "
                        "veya farklı bir isim girin."
                    )

                amb_secenekleri = ["Ambalaj", "Tank", "Dökme"] if v2 else ["AMBALAJLI", "TANK", "DÖKME"]
                urun["ambalaj"] = st.selectbox(
                    "Ambalajlı / Tank / Dökme",
                    amb_secenekleri,
                    index=amb_secenekleri.index(urun["ambalaj"]) if urun["ambalaj"] in amb_secenekleri else 0,
                    key=f"amb_{fname}",
                )

                urun["dahil_et"] = st.checkbox("Bu ürünü Excel'e eklemeye dahil et", value=urun["dahil_et"], key=f"dahil_{fname}")

            with col2:
                st.markdown("**Bölüm 14'ten okunan:**")
                st.write(f"- UN No: `{info.get('un_no') or '-'}`")
                st.write(f"- Sınıf: `{info.get('sinif') or '-'}`")
                st.write(f"- Paketleme Grubu: `{info.get('paketleme_grubu') or '-'}`")
                st.write(f"- Revize Tarihi: `{info.get('revize_tarihi') or '-'}`")

            if v2:
                row_preview = build_inventory_row_v2(info, urun["kimyasal_adi"], urun["ambalaj"])
            else:
                row_preview = build_inventory_row(info, TABLO_A_PATH, urun["kimyasal_adi"], urun["ambalaj"])

            if row_preview["durum"] == "ok":
                if v2:
                    st.markdown("**Bölüm 14'ten yazılacak veriler:**")
                    st.table({k: [v] for k, v in row_preview.items()
                              if k not in ("Kimyasal Adı", "Ambalaj/Tank", "MSDS/SDS TARİHİ", "durum")})
                else:
                    st.markdown("**Tablo A'dan eşleşen veriler:**")
                    st.table({k: [v] for k, v in row_preview.items()
                              if k not in ("Kimyasal Adı", "AMBALAJLI/TANK/DÖKME", "MSDS/SDS TARİHİ", "durum")})
            elif row_preview["durum"] == "manual_review":
                st.error("Bu satır 'MANUEL KONTROL GEREKLİ' olarak işaretlenecek. "
                         "Excel'e aktarıldıktan sonra ilgili hücreleri elle düzeltin.")
            else:
                st.info("Bu ürün SDS Bölüm 14'e göre ADR kapsamında değil; ilgili sabit metin yazılacak.")

    st.divider()
    secili_urunler = [u for u in st.session_state.urunler.values() if u["dahil_et"]]
    render_export_ui(secili_urunler, envanter_path, v2, firma_adi, key_suffix="bottom")
elif not envanter_path:
    st.info("Önce sol menüden envanter Excel dosyasını yükleyin.")
else:
    st.info("İşlenecek MSDS PDF dosyalarını yükleyin.")

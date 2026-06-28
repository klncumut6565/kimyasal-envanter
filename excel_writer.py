"""
Envanter Excel dosyasına (ASUTEK ... rev1.xlsx formatı) yeni ürün satırları
ekleyen modül. Kurallar:
  - Yeni satırlar, alt taraftaki "KONTROL EDEN / ONAYLAYAN" imza bloğunun
    ÜZERİNE, son ürün satırından hemen sonra eklenir.
  - Her hücre, kendisinden bir önceki (üstündeki) satırdaki hücrenin
    biçimini (font, kenarlık, hizalama, wrap, satır yüksekliği) birebir alır.
  - "No" sütunu otomatik artar.
  - Boş "Y" sütunu "Logo" başlığıyla kullanılır; ürün logosu resim olarak
    bu hücreye gömülür.
"""
import copy
import os
import re
import datetime
import difflib
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter

from extractor import clean_product_name

SHEET_NAME = "Tesis_Genelinde_Kull._ İnditex"
SHEET_NAME_V2 = "KİMYASAL ENVANTER2026"  # Versiyon 2 dosyalarındaki sayfa adı
HEADER_ROW = 5
STYLE_TEMPLATE_ROW = HEADER_ROW + 1  # hiç ürün yokken kullanılacak biçim satırı
FOOTER_MARKER = "KONTROL EDEN"

# Versiyon 2'de bu sütun(lar)a hiçbir zaman otomatik yazılmaz -- elle
# atanan, MSDS Bölüm 14'ten gelmeyen organizasyonel bilgilerdir.
V2_PROTECTED_COLUMNS = {"Kullanılan Bölüm"}


def _norm(s):
    """Boşluk/satır-sonu farklarına duyarsız karşılaştırma için normalize eder
    (örn. 'AMBALAJLI/TANK/DÖKME' ile 'AMBALAJLI/ TANK/ DÖKME' eşleşir)."""
    return re.sub(r"\s+", "", str(s)).upper() if s else ""


def build_column_map(ws):
    """Başlık satırını okuyarak {normalize edilmiş başlık: sütun no} çıkarır.
    Sütun sırası değişse/yeni sütun eklense bile kod kırılmaz."""
    mapping = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=HEADER_ROW, column=c).value
        if isinstance(v, str) and v.strip():
            mapping[_norm(v)] = c
    return mapping


def find_column(ws, header_text):
    """Tek bir başlık adının sütun numarasını döner, yoksa None."""
    return build_column_map(ws).get(_norm(header_text))


def _title_cell(ws, row):
    """1. veya 3. satırdaki başlık birleştirme aralığının (firma logosu
    hariç) sol üst hücresini bulur. Sütunlar yeniden sıralanıp bu
    birleştirme kaysa bile doğru hücreyi bulmaya devam eder."""
    for m in ws.merged_cells.ranges:
        if m.min_row == row and m.min_col > 2:
            return ws.cell(row=m.min_row, column=m.min_col)
    return None


def _set_value_right_of_label(ws, label_prefix, value, max_row=4):
    """'Yayın Tarihi:' gibi bir etiket hücresinin SAĞINDAKİ hücreye değer
    yazar. Etiket hangi sütunda olursa olsun çalışır (sütun kaymalarına
    dayanıklı)."""
    for row in ws.iter_rows(min_row=1, max_row=max_row):
        for cell in row:
            if isinstance(cell.value, str) and cell.value.strip().startswith(label_prefix):
                ws.cell(row=cell.row, column=cell.column + 1, value=value)
                return True
    return False


def _find_footer_row(ws):
    for r in range(HEADER_ROW + 1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=r, column=c).value
            if isinstance(v, str) and FOOTER_MARKER in v:
                return r
    return None


def _find_last_data_row(ws, footer_row):
    for r in range(footer_row - 1, HEADER_ROW, -1):
        if ws.cell(row=r, column=1).value is not None:
            return r
    return HEADER_ROW


def _strip_bold(ws, row, max_col):
    """Verilen satırdaki hiçbir hücre kalın (bold) yazılmasın -- hangi
    kaynak satırdan stil kopyalanmış olursa olsun (yeni şablon veya var
    olan dosyadaki önceki satır), eklenen ürün satırları sayfa genelinde
    hiçbir zaman kalın görünmeyecek."""
    for c in range(1, max_col + 1):
        cell = ws.cell(row=row, column=c)
        if cell.font.bold:
            f = copy.copy(cell.font)
            f.bold = False
            cell.font = f


def _estimate_lines(text, col_width, font_size):
    """Bir metnin, verilen sütun genişliği ve font boyutunda kaç satıra
    sarılacağını kabaca tahmin eder (openpyxl 'width' birimi ~11 punto
    Calibri karakter genişliğine denktir)."""
    if not text:
        return 1
    chars_per_line = max(1, int((col_width or 10) * (11.0 / max(font_size or 11, 6))))
    total = 0
    for part in str(text).split("\n"):
        total += 1 if part == "" else -(-len(part) // chars_per_line)  # ceil
    return max(total, 1)


def _auto_fit_row_height(ws, row, max_col, min_height):
    """Hücre içeriği (özellikle uzun 'kapsam dışı' / 'manuel kontrol'
    metinleri) mevcut satır yüksekliğine sığmıyorsa, satırı otomatik
    olarak uzatır. Kısa içerikte şablonun standart yüksekliği korunur,
    hücre taşması/yazının kesilmesi engellenir."""
    max_lines = 1
    for c in range(1, max_col + 1):
        cell = ws.cell(row=row, column=c)
        if not isinstance(cell.value, str) or not cell.alignment.wrap_text:
            continue
        col_letter = get_column_letter(c)
        width = ws.column_dimensions[col_letter].width
        lines = _estimate_lines(cell.value, width, cell.font.size)
        max_lines = max(max_lines, lines)
    needed = max_lines * 22  # ~satır başına nokta (büyük fontlara pay bırakır)
    if needed > (min_height or 0):
        ws.row_dimensions[row].height = needed
    elif min_height:
        ws.row_dimensions[row].height = min_height


def _copy_row_style(ws, src_row, dst_row, max_col):
    for c in range(1, max_col + 1):
        src = ws.cell(row=src_row, column=c)
        dst = ws.cell(row=dst_row, column=c)
        dst.font = copy.copy(src.font)
        dst.border = copy.copy(src.border)
        dst.fill = copy.copy(src.fill)
        dst.alignment = copy.copy(src.alignment)
        dst.number_format = src.number_format
        dst.protection = copy.copy(src.protection)
    if ws.row_dimensions[src_row].height is not None:
        ws.row_dimensions[dst_row].height = ws.row_dimensions[src_row].height


def ensure_note_header(ws):
    """'Açıklama' başlıklı sütunu bulur; yoksa son sütunun hemen sağına
    ekler (önceden 'Logo' idi -- artık piktogram/sınıf tutarsızlığı gibi
    uyarı metinleri için kullanılıyor, resim için değil).
    Döndürdüğü sütun numarası, ileride uyarı metni yazarken kullanılır."""
    col_map = build_column_map(ws)
    note_col = col_map.get(_norm("Açıklama"))
    if note_col:
        return note_col

    last_col = max(col_map.values()) if col_map else ws.max_column
    note_col = last_col + 1
    donor = ws.cell(row=HEADER_ROW, column=last_col)
    cell = ws.cell(row=HEADER_ROW, column=note_col, value="Açıklama")
    cell.font = copy.copy(donor.font)
    cell.alignment = copy.copy(donor.alignment)
    cell.border = copy.copy(donor.border)
    cell.fill = copy.copy(donor.fill)
    col_letter = get_column_letter(note_col)
    if ws.column_dimensions[col_letter].width in (None, 0):
        ws.column_dimensions[col_letter].width = 90
    return note_col


def _shift_rows_down(ws, start_row, end_row, n, max_col):
    """[start_row, end_row] satır aralığını n satır aşağı kaydırır.

    NOT: ws.insert_rows() kullanmıyoruz çünkü openpyxl bu fonksiyonda
    birleştirilmiş (merged) hücreleri doğru taşımıyor; bu da alt
    taraftaki "KONTROL EDEN/ONAYLAYAN" imza bloğunun içeriğinin
    kaybolmasına neden oluyordu. Bu yüzden hücreleri ve birleştirme
    aralıklarını elle, alttan üste doğru kaydırıyoruz.
    """
    affected_merges = [m for m in list(ws.merged_cells.ranges) if m.min_row >= start_row]
    for m in affected_merges:
        ws.unmerge_cells(str(m))

    for r in range(end_row, start_row - 1, -1):
        dst = r + n
        for c in range(1, max_col + 1):
            src_cell = ws.cell(row=r, column=c)
            dst_cell = ws.cell(row=dst, column=c)
            dst_cell.value = src_cell.value
            dst_cell.font = copy.copy(src_cell.font)
            dst_cell.border = copy.copy(src_cell.border)
            dst_cell.fill = copy.copy(src_cell.fill)
            dst_cell.alignment = copy.copy(src_cell.alignment)
            dst_cell.number_format = src_cell.number_format
            dst_cell.protection = copy.copy(src_cell.protection)
        if ws.row_dimensions[r].height is not None:
            ws.row_dimensions[dst].height = ws.row_dimensions[r].height

    for m in affected_merges:
        ws.merge_cells(start_row=m.min_row + n, end_row=m.max_row + n,
                        start_column=m.min_col, end_column=m.max_col)

    # Eski (şimdi kopyalanmış) satırları temizle, böylece yeni ürün
    # satırları için boş alan açılmış olur.
    for r in range(start_row, start_row + n):
        for c in range(1, max_col + 1):
            ws.cell(row=r, column=c).value = None


def create_new_envanter(template_path: str, output_path: str, firma_adi: str,
                          logo_path: str = None, hazirlayan_adi: str = None,
                          onaylayan_adi: str = None):
    """
    Boş şablon (data/BOS_ENVANTER_SABLONU.xlsx) üzerinden, firma adı,
    (varsa) firma logosu, hazırlayan adı ve onaylayan adı işlenmiş, ürün
    eklemeye hazır yeni bir envanter dosyası oluşturur. Excel hiç
    yüklenmediğinde app.py bunu kullanır.
    """
    wb = load_workbook(template_path)
    ws = wb[SHEET_NAME]

    firma_adi = (firma_adi or "").strip()
    title_cell = _title_cell(ws, 1)
    if title_cell is not None:
        title_cell.value = f"{firma_adi} TAKİP LİSTESİ(ENVANTER)".strip()
    bugun = datetime.date.today()
    _set_value_right_of_label(ws, "Yayın Tarihi", bugun)
    _set_value_right_of_label(ws, "Revizyon Tarihi", bugun)

    hazirlayan_adi = (hazirlayan_adi or "").strip()
    if hazirlayan_adi:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("HAZIRLAYAN"):
                    cell.value = (
                        f"HAZIRLAYAN\n{hazirlayan_adi}\nTehlikeli Madde Güvenlik Danışmanı"
                    )
                    break

    onaylayan_adi = (onaylayan_adi or "").strip()
    if onaylayan_adi:
        # Bilgi girilmişse "Sorumlu Kişi" yerine girilen ad-soyad yazılır.
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("ONAYLAYAN"):
                    cell.value = f"ONAYLAYAN\n{onaylayan_adi}"
                    break
    # onaylayan_adi boşsa hiçbir şey değiştirilmez, şablondaki varsayılan
    # "ONAYLAYAN\nSorumlu Kişi" metni olduğu gibi kalır.

    if logo_path and os.path.exists(logo_path):
        img = XLImage(logo_path)
        row_h = sum(ws.row_dimensions[r].height or 20 for r in range(1, 5))
        col_w = sum(ws.column_dimensions[c].width or 10 for c in ("A", "B"))
        target_h_px = row_h * 1.33
        target_w_px = col_w * 7
        scale = min(target_h_px / img.height, target_w_px / img.width)
        img.height = int(img.height * scale)
        img.width = int(img.width * scale)
        ws.add_image(img, "A1")

    wb.save(output_path)
    return output_path


def _normalize_name(s):
    """Ürün adlarını karşılaştırmak için normalize eder: dosya adı ekleri
    (MSDS, rev 7, CLP Türkçe Türkiye 12292025 vb.) temizlenir, NBSP/çoklu
    boşluk farklarına duyarsız hale getirilir, büyük harfe çevrilir."""
    return clean_product_name(s).upper()


def _find_existing_row_by_name(ws, kimyasal_adi, name_col, header_row, last_row):
    """Önce tam eşleşmeyi dener (temizlenmiş isimle); bulamazsa, biri
    diğerini kapsıyorsa veya yazım hatasına toleranslı benzerlik oranı
    yüksekse (typo'lar -- örn. 'AMETHSYT' / 'AMETHYST') en iyi adayı
    bulanık eşleştirme ile bulur. Çok kısa/alakasız isimlerde yanlış
    eşleşmeyi önlemek için eşik değeri korunur.
    """
    target = _normalize_name(kimyasal_adi)
    if not target:
        return None

    best_row, best_score = None, 0.0
    for r in range(header_row + 1, last_row + 1):
        cell_val = _normalize_name(ws.cell(row=r, column=name_col).value)
        if not cell_val:
            continue
        if cell_val == target:
            return r  # tam eşleşme (temizlenmiş), kesin doğru -- hemen dön

        score = difflib.SequenceMatcher(None, cell_val, target).ratio()
        shorter, longer = (cell_val, target) if len(cell_val) <= len(target) else (target, cell_val)
        if shorter and shorter in longer:
            score = max(score, len(shorter) / len(longer) * 0.97)

        if score > best_score:
            best_score, best_row = score, r

    return best_row if best_score >= 0.84 else None


_PICTOGRAM_COLUMNS_V2 = {"Tehlike_Etiketi", "ADR_İŞARETİ"}


def _satirda_gercek_un_no_var_mi(ws, row, un_col):
    """Hücredeki mevcut değerin gerçek bir UN numarası olup olmadığını
    kontrol eder ('UN 2014', 'UN2014' veya sadece '2014' gibi). Sabit
    metinler (MANUEL KONTROL / kapsam dışı) bu deseni karşılamaz."""
    if not un_col:
        return False
    val = ws.cell(row=row, column=un_col).value
    if not val:
        return False
    return bool(re.match(r"^\s*(UN\s*)?\d{3,4}\s*$", str(val), re.IGNORECASE))


def fill_or_append_v2(envanter_path: str, output_path: str, urunler: list):
    """Versiyon 2 (ORDU tarzı basit format) için: ürün adı zaten bir
    satırda varsa ve ADR hücreleri BOŞSA onları doldurur (dolu hücrelere
    dokunmaz). Ürün adı envanterde bulunamazsa o ürün ATLANIR -- yeni
    satır eklenmez (Versiyon 2'nin amacı sadece var olan boş hücreleri
    tamamlamaktır, yeni ürün eklemek değildir).

    Dönüş: [(satır_no veya None, "dolduruldu"|"zaten_dolu"|"eslesme_yok", kimyasal_adi), ...]
    """
    wb = load_workbook(envanter_path)
    if SHEET_NAME_V2 not in wb.sheetnames:
        raise ValueError(
            f"Yüklenen dosyada '{SHEET_NAME_V2}' adlı bir sayfa bulunamadı. "
            f"Dosyadaki sayfalar: {', '.join(wb.sheetnames)}"
        )
    ws = wb[SHEET_NAME_V2]
    max_col = ws.max_column

    col_map = build_column_map(ws)
    name_col = col_map.get(_norm("Kimyasal Adı"))
    if not name_col:
        raise ValueError("'Kimyasal Adı' sütunu bulunamadı; dosya yapısı beklenenden farklı.")

    last_row = ws.max_row
    while last_row > HEADER_ROW and all(
        ws.cell(row=last_row, column=c).value in (None, "") for c in range(1, max_col + 1)
    ):
        last_row -= 1

    sonuc = []
    for urun in urunler:
        ad = urun.get("Kimyasal Adı")
        existing_row = _find_existing_row_by_name(ws, ad, name_col, HEADER_ROW, last_row)

        if not existing_row:
            sonuc.append((None, "eslesme_yok", ad))
            continue

        herhangi_dolduruldu = False
        un_col = col_map.get(_norm("UN NUMARASI"))
        # "Bu üründe UN no var" durumunu hem bu işlemdeki yeni veriden hem de
        # satırın ZATEN sahip olduğu mevcut değerden kontrol ediyoruz --
        # aksi halde, satır önceden (başka bir eşleşmeden) gerçek bir UN no
        # almışken, bu seferki "kapsam dışı" sonucu ADR_İŞARETİ/Tehlike_Etiketi'ni
        # yanlışlıkla doldurabilir.
        yeni_un_str = str(urun.get("UN NUMARASI") or "")
        bu_islemde_un_var = bool(re.match(r"^\s*UN\s*\d{3,4}\s*$", yeni_un_str, re.IGNORECASE))
        un_no_var = bu_islemde_un_var or _satirda_gercek_un_no_var_mi(ws, existing_row, un_col)

        for col_name, value in urun.items():
            if value in (None, ""):
                continue  # PDF'ten bu alan çıkarılamadıysa hücreye dokunma
            if col_name in ("Kimyasal Adı", "durum") or _norm(col_name) in {_norm(c) for c in V2_PROTECTED_COLUMNS}:
                continue
            if un_no_var and _norm(col_name) in {_norm(c) for c in _PICTOGRAM_COLUMNS_V2}:
                continue  # UN no varsa bu iki sütuna asla metin yazılmaz (piktogram elle eklenecek)
            idx = col_map.get(_norm(col_name))
            if idx and ws.cell(row=existing_row, column=idx).value in (None, ""):
                ws.cell(row=existing_row, column=idx, value=value)
                herhangi_dolduruldu = True

        if herhangi_dolduruldu:
            _strip_bold(ws, existing_row, max_col)
            # NOT: Burada satır yüksekliğine kasıtlı olarak dokunulmuyor --
            # bu, var olan bir satırı dolduruyoruz (Versiyon 2), yüksekliği
            # zaten kullanıcının kendi tasarımının bir parçası. Sadece
            # YENİ satır eklerken (add_products / V1) otomatik sığdırma
            # yapılır, çünkü o durumda satır sıfırdan oluşturuluyor.
            sonuc.append((existing_row, "dolduruldu", ad))
        else:
            sonuc.append((existing_row, "zaten_dolu", ad))

    wb.save(output_path)
    return sonuc


def add_products(envanter_path: str, output_path: str, urunler: list):
    """
    urunler: matcher.build_inventory_row() çıktısı sözlüklerin listesi.
             Her sözlükte opsiyonel "logo_path" anahtarı olabilir
             (yerel diskteki resim dosyası yolu).
    Dönen değer: eklenen satırların Excel satır numaraları listesi.
    """
    wb = load_workbook(envanter_path)
    ws = wb[SHEET_NAME]

    footer_row = _find_footer_row(ws)
    if footer_row is None:
        raise ValueError("KONTROL EDEN / ONAYLAYAN imza bloğu bulunamadı; "
                          "dosya yapısı değişmiş olabilir.")
    last_data_row = _find_last_data_row(ws, footer_row)
    # Henüz hiç ürün satırı yoksa (boş şablon), "No" 0'dan başlar ve ilk
    # ürünün biçimi başlığın altındaki özel stil şablonu satırından alınır
    # (aksi halde başlık satırının biçimi -mavi, kalın- yanlışlıkla kopyalanır).
    bos_sablon = (last_data_row == HEADER_ROW)
    last_no = 0 if bos_sablon else (ws.cell(row=last_data_row, column=1).value or 0)

    note_col = ensure_note_header(ws)
    col_map = build_column_map(ws)  # ensure_note_header sonrası tekrar oku
    max_col = max(ws.max_column, note_col)
    no_col = col_map.get(_norm("No"), 1)

    n = len(urunler)
    insert_at = last_data_row + 1
    _shift_rows_down(ws, footer_row, ws.max_row, n, max_col)

    added_rows = []
    for i, urun in enumerate(urunler):
        target_row = insert_at + i
        style_src_row = (STYLE_TEMPLATE_ROW if bos_sablon else last_data_row) if i == 0 else target_row - 1
        _copy_row_style(ws, style_src_row, target_row, max_col)
        _strip_bold(ws, target_row, max_col)

        ws.cell(row=target_row, column=no_col, value=int(last_no) + i + 1)
        for col_name, value in urun.items():
            idx = col_map.get(_norm(col_name))
            if idx and idx != no_col:
                ws.cell(row=target_row, column=idx, value=value)

        _auto_fit_row_height(ws, target_row, max_col, ws.row_dimensions[target_row].height)

        added_rows.append(target_row)

    # Üst bilgi alanındaki "Revizyon Tarihi", dosya her güncellendiğinde
    # otomatik olarak bugünün tarihine güncellenir.
    _set_value_right_of_label(ws, "Revizyon Tarihi", datetime.date.today())

    wb.save(output_path)
    return added_rows


if __name__ == "__main__":
    import sys
    sys.path.insert(0, ".")
    from extractor import extract_adr_info
    from matcher import build_inventory_row

    tablo_a = "ADR_A_TABLOSU.xlsx"
    envanter = "ASUTEK_Kimyasal_İnceleme_Kimyasal_Envanter__ADR_rev1.xlsx"
    out = "TEST_OUTPUT_envanter.xlsx"

    urunler = []
    for pdf, ad in [
        ("ADVANTAGE_101M_MSDS_TÜRKÇE.PDF", "ADVANTAGE 101M (TEST)"),
        ("ADVANTAGE_121_ODT_MSDS_TÜRKÇE.pdf", "ADVANTAGE 121 ODT (TEST)"),
    ]:
        info = extract_adr_info(pdf)
        row = build_inventory_row(info, tablo_a, ad)
        urunler.append(row)

    added = add_products(envanter, out, urunler)
    print("Eklenen satırlar:", added)
